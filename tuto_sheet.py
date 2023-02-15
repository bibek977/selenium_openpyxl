import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

workbook = openpyxl.Workbook()
sheet = workbook.active

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
driver.get("https://rahulshettyacademy.com/AutomationPractice/")

sheet.title = "Demo Sheet"

# table_data = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']/table[@id='product']/tbody/tr")
table_row = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr")

i = 1
for data in table_row:

    sheet.cell(i, 1).value = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[1]").text
    sheet.cell(i, 2).value = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[2]").text
    sheet.cell(i, 3).value = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[3]").text
    sheet.cell(i, 4).value = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr["+str(i)+"]/td[4]").text

    i += 1

workbook.save("Excel_sheets/Demo_1.xlsx")