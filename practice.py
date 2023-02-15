import openpyxl

workbook = openpyxl.load_workbook("Excel_sheets/bills.xlsx")

sheet = workbook['Sheet1']

a = sheet['A1'].value
b = sheet.cell(1, 1).value
print(a,b)

c = sheet['D7'].value
d = sheet.cell(7,4).value
print(c,d)